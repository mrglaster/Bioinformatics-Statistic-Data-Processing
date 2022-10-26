import os
import numpy as np
import openpyxl
import statistics
import re
import magic

from openpyxl import load_workbook
from google.colab import files
from zipfile import ZipFile

result_filenames = []
DEBUG = 0

def process_matrix(arrDataMatrix):
    """Работа с матрицей с n строками и k столбцами. На выходе - exel файл"""
    nRows = len(arrDataMatrix)
    nCols = len(arrDataMatrix[0])

    if nCols == 0 or nRows == 0:
      print(arrDataMatrix)
      raise ValueError(f"Unable to process matrix. Incorrect size: {nRows}x{nCols}")

    arrMidCap = list(np.zeros(shape=nCols, dtype=float))
    arrStCap = list(np.zeros(shape=nCols, dtype=float))
    fExelBook = openpyxl.Workbook()


    shSourceSheet = fExelBook.create_sheet("SourceData")
    shDataSheet = fExelBook.create_sheet("ModifiedMatrix")
    std = fExelBook.get_sheet_by_name('Sheet')
    fExelBook.remove_sheet(std)

    for i in range(nRows):
        for j in range(nCols):
            fCurNumber = arrDataMatrix[i][j]
            shSourceSheet.cell(i + 1, j + 1).value=fCurNumber

    arrDataMatrixTransposed = np.transpose(arrDataMatrix)
    for i in range(len(arrDataMatrixTransposed)):
        arrMidCap[i] = sum(list(arrDataMatrixTransposed[i])) / nRows
        arrStCap[i] = statistics.stdev(list(arrDataMatrixTransposed[i]))

    for i in range(nRows):
        for j in range(nCols):
            result_number = (arrDataMatrix[i][j] - arrMidCap[j]) / arrStCap[j]
            shDataSheet.cell(i + 1, j + 1).value = result_number
    filename = "ExtractsTables_len_"+str(nCols)+".xlsx"
    fExelBook.save(filename)
    result_filenames.append(filename)
    #files.download('ExtractsTables.xlsx')


def remove_spaces(sLine):
  """Удаляет лишние пробелы из строки"""
  if DEBUG: 
    print(f"Current line pre: {sLine}")

  current_line = current_line = re.sub(" +", " ", sLine).replace(",",".").replace('\t', ' ').split(' ')
  
  if DEBUG:
    print(f"Current line post: {current_line}")

  return current_line[:len(current_line)-1]

def get_encoding(filename):
  """Получить кодировку файла"""
  blob = open(filename, 'rb').read()
  m = magic.Magic(mime_encoding=True)
  encoding = m.from_buffer(blob)
  return encoding


def load_data(sInputFile):
  """Загружает данные из текстового файла в матрицу"""
  with open(sInputFile, "r", encoding=get_encoding(sInputFile)) as f:
      arrDataMatrix = []
      for line in f:
        current_line = remove_spaces(line)
        try:
          arrDataMatrix.append(np.array(current_line).astype(np.float))

        except:
          pass
  return arrDataMatrix


def download_files():
  """Скачивает все .xlsx файлы, полученные в процессе работы process_matrix"""
  global result_filenames

  # Если только один файл
  if len(result_filenames) == 1:
    files.download(result_filenames[0])

  # Если больше одного файла
  elif len(result_filenames) > 1: 
    archive_name = "results.zip"
    zipObj = ZipFile(archive_name, 'w')
    for i in result_filenames:
      zipObj.write(i)
    zipObj.close()
    files.download(archive_name)

  # Если что-то пошло не так
  else:
    raise ValueError("Unable to download results: no results data")

def split_matrix(arrDataMatrix):
  """В случае, если в исходной матрице есть строки разных длин - разделяет её на несколько матриц"""
  if len(arrDataMatrix) == 0:
    raise ValueError("Unable to split matrix: input data is empty!")
  submatrixes = {}
  for i in arrDataMatrix:
    iCurrentLen = len(i)
    if iCurrentLen not in submatrixes.keys():
      submatrixes[iCurrentLen] = []
    submatrixes[iCurrentLen].append(i)
  if len(submatrixes.keys()) != 1:
    print(f"Source matrix was separated to {len(submatrixes.keys())} matrixes with different lengths")
  return submatrixes


def process_all(submatrixes):
  """Обработать все матрицы из словаря"""
  if len(submatrixes.keys()) == 0:
    raise ValueError("Unable to process submatrixes: submatrixes not found")
  for i in submatrixes.keys():
    process_matrix(submatrixes[i])




def main():
    """Главная функция"""

    #Загрузка файла в Google Colaboratory
    uploaded = files.upload()
    sInputFile = next(iter(uploaded))

    #Подготовка данных
    arrDataMatrix = load_data(sInputFile)
    submatrixes = split_matrix(arrDataMatrix) 
    
    #Обработка данных
    process_all(submatrixes)

    #Скачивание .xlsx таблиц с результатами    
    download_files()


if __name__ == '__main__':
  main()
